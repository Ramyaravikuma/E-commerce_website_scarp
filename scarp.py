import requests 
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook
wb = Workbook()
ws1 = wb.create_sheet("Sheet_A")
ws1.title = "covid essential"
ws2 = wb.create_sheet("Sheet_B", 0)
ws2.title = "diabetes"
ws3 = wb.create_sheet("Sheet_c", 1)
ws3.title = "eyewear"
ws4 = wb.create_sheet("Sheet_c", 1)
ws4.title = "ayush"

ws1.append(["product Name","Prize(Rs)","Dealer"])
ws2.append(["product Name","Prize(Rs)","Dealer"])
ws3.append(["product Name","Prize(Rs)","Dealer"])
ws4.append(["product Name","Prize(Rs)","Dealer"])
#covid essential
try:
    
    link=["https://www.netmeds.com/non-prescriptions/covid-essentials","https://www.netmeds.com/non-prescriptions/covid-essentials/page/2",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/3","https://www.netmeds.com/non-prescriptions/covid-essentials/page/4",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/5",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/6","https://www.netmeds.com/non-prescriptions/covid-essentials/page/7",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/8",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/9","https://www.netmeds.com/non-prescriptions/covid-essentials/page/10",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/11","https://www.netmeds.com/non-prescriptions/covid-essentials/page/12",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/13",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/14","https://www.netmeds.com/non-prescriptions/covid-essentials/page/15",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/16",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/17",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/18","https://www.netmeds.com/non-prescriptions/covid-essentials/page/19",
        "https://www.netmeds.com/non-prescriptions/covid-essentials/page/20"]
    for links in link:
        html = requests.get(links) 
        #html = links.content
        soup = BeautifulSoup(html.text, 'html.parser')
        #print(soup)
        product=soup.find("div",class_="row product-list").find_all("div",class_="cat-item")
        for products in product:
            #print(products)
            product_name=products.find("a").find("span",class_="clsgetname").get_text(strip=True)
            product_price=products.find('span', attrs={'id':'final_price'}).get_text(strip=True).replace("Rs.","")
            product_owner=products.find("span",class_="drug-varients ellipsis").get_text(strip=True).replace("Mkt:","")
            #print(product_name,product_price,product_owner)
            ws1.append([product_name,product_price,product_owner])
    
    
except Exception as e:
    print(e)

#diabetes
try:
    
    link=["https://www.netmeds.com/non-prescriptions/diabetes-support","https://www.netmeds.com/non-prescriptions/diabetes-support/page/2",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/3","https://www.netmeds.com/non-prescriptions/diabetes-support/page/4",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/5","https://www.netmeds.com/non-prescriptions/diabetes-support/page/6",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/7","https://www.netmeds.com/non-prescriptions/diabetes-support/page/8",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/9","https://www.netmeds.com/non-prescriptions/diabetes-support/page/10",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/11","https://www.netmeds.com/non-prescriptions/diabetes-support/page/12",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/13","https://www.netmeds.com/non-prescriptions/diabetes-support/page/14",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/15","https://www.netmeds.com/non-prescriptions/diabetes-support/page/16",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/17","https://www.netmeds.com/non-prescriptions/diabetes-support/page/18",
            "https://www.netmeds.com/non-prescriptions/diabetes-support/page/19","https://www.netmeds.com/non-prescriptions/diabetes-support/page/20"]
    for links in link:
        html = requests.get(links) 
        #html = links.content
        soup = BeautifulSoup(html.text, 'html.parser')
        #print(soup)
        product=soup.find("div",class_="row product-list").find_all("div",class_="cat-item")
        for products in product:
            #print(products)
            product_name=products.find("a").find("span",class_="clsgetname").get_text(strip=True)
            product_price=products.find('span', attrs={'id':'final_price'}).get_text(strip=True).replace("Rs.","")
            product_owner=products.find("span",class_="drug-varients ellipsis").get_text(strip=True).replace("Mkt:","")
            #print(product_name,product_price,product_owner)
            ws2.append([product_name,product_price,product_owner])
            
            
except Exception as e:
    print(e)
    
#eye wear        
try:
    
    link=["https://www.netmeds.com/non-prescriptions/eyewear","https://www.netmeds.com/non-prescriptions/eyewear/page/2",
          "https://www.netmeds.com/non-prescriptions/eyewear/page/3","https://www.netmeds.com/non-prescriptions/eyewear/page/4",
          "https://www.netmeds.com/non-prescriptions/eyewear/page/5","https://www.netmeds.com/non-prescriptions/eyewear/page/6",
          "https://www.netmeds.com/non-prescriptions/eyewear/page/7","https://www.netmeds.com/non-prescriptions/eyewear/page/8",
          "https://www.netmeds.com/non-prescriptions/eyewear/page/9","https://www.netmeds.com/non-prescriptions/eyewear/page/10",
          "https://www.netmeds.com/non-prescriptions/eyewear/page/11"]
    
    for links in link:
        html = requests.get(links) 
        #html = links.content
        soup = BeautifulSoup(html.text, 'html.parser')
        #print(soup)
        product=soup.find("div",class_="row product-list").find_all("div",class_="cat-item")
        for products in product:
            #print(products)
            product_name=products.find("a").find("span",class_="clsgetname").get_text(strip=True)
            product_price=products.find('span', attrs={'id':'final_price'}).get_text(strip=True).replace("Rs.","")
            product_owner=products.find("span",class_="drug-varients ellipsis").get_text(strip=True).replace("Mkt:","")
            #print(product_name,product_price,product_owner)
            ws3.append([product_name,product_price,product_owner])
    
    
except Exception as e:
    print(e)

# ayush     
try:
    
    link=["https://www.netmeds.com/non-prescriptions/ayush","https://www.netmeds.com/non-prescriptions/ayush/page/2",
          "https://www.netmeds.com/non-prescriptions/ayush/page/3","https://www.netmeds.com/non-prescriptions/ayush/page/4",
          "https://www.netmeds.com/non-prescriptions/ayush/page/6","https://www.netmeds.com/non-prescriptions/ayush/page/7",
          "https://www.netmeds.com/non-prescriptions/ayush/page/8","https://www.netmeds.com/non-prescriptions/ayush/page/9",
          "https://www.netmeds.com/non-prescriptions/ayush/page/10","https://www.netmeds.com/non-prescriptions/ayush/page/11",
          "https://www.netmeds.com/non-prescriptions/ayush/page/12","https://www.netmeds.com/non-prescriptions/ayush/page/13",
          "https://www.netmeds.com/non-prescriptions/ayush/page/14","https://www.netmeds.com/non-prescriptions/ayush/page/15",
          "https://www.netmeds.com/non-prescriptions/ayush/page/16","https://www.netmeds.com/non-prescriptions/ayush/page/5",
          "https://www.netmeds.com/non-prescriptions/ayush/page/17","https://www.netmeds.com/non-prescriptions/ayush/page/18",
          "https://www.netmeds.com/non-prescriptions/ayush/page/19","https://www.netmeds.com/non-prescriptions/ayush/page/20",
          "https://www.netmeds.com/non-prescriptions/ayush/page/21","https://www.netmeds.com/non-prescriptions/ayush/page/22",
          "https://www.netmeds.com/non-prescriptions/ayush/page/23","https://www.netmeds.com/non-prescriptions/ayush/page/24",
          "https://www.netmeds.com/non-prescriptions/ayush/page/25","https://www.netmeds.com/non-prescriptions/ayush/page/26",
          "https://www.netmeds.com/non-prescriptions/ayush/page/27","https://www.netmeds.com/non-prescriptions/ayush/page/28",
          "https://www.netmeds.com/non-prescriptions/ayush/page/29","https://www.netmeds.com/non-prescriptions/ayush/page/30",
          "https://www.netmeds.com/non-prescriptions/ayush/page/31","https://www.netmeds.com/non-prescriptions/ayush/page/32",
          "https://www.netmeds.com/non-prescriptions/ayush/page/33","https://www.netmeds.com/non-prescriptions/ayush/page/34",
          "https://www.netmeds.com/non-prescriptions/ayush/page/35","https://www.netmeds.com/non-prescriptions/ayush/page/36",
          "https://www.netmeds.com/non-prescriptions/ayush/page/37","https://www.netmeds.com/non-prescriptions/ayush/page/38",
          "https://www.netmeds.com/non-prescriptions/ayush/page/39","https://www.netmeds.com/non-prescriptions/ayush/page/40",
          "https://www.netmeds.com/non-prescriptions/ayush/page/41","https://www.netmeds.com/non-prescriptions/ayush/page/42",
          "https://www.netmeds.com/non-prescriptions/ayush/page/43","https://www.netmeds.com/non-prescriptions/ayush/page/44",
          "https://www.netmeds.com/non-prescriptions/ayush/page/45","https://www.netmeds.com/non-prescriptions/ayush/page/46",
          "https://www.netmeds.com/non-prescriptions/ayush/page/47","https://www.netmeds.com/non-prescriptions/ayush/page/48",
          "https://www.netmeds.com/non-prescriptions/ayush/page/49","https://www.netmeds.com/non-prescriptions/ayush/page/50",
          "https://www.netmeds.com/non-prescriptions/ayush/page/51","https://www.netmeds.com/non-prescriptions/ayush/page/52",
          "https://www.netmeds.com/non-prescriptions/ayush/page/53","https://www.netmeds.com/non-prescriptions/ayush/page/54",
          "https://www.netmeds.com/non-prescriptions/ayush/page/55","https://www.netmeds.com/non-prescriptions/ayush/page/56",
          "https://www.netmeds.com/non-prescriptions/ayush/page/57","https://www.netmeds.com/non-prescriptions/ayush/page/58",
          "https://www.netmeds.com/non-prescriptions/ayush/page/59","https://www.netmeds.com/non-prescriptions/ayush/page/60",
          "https://www.netmeds.com/non-prescriptions/ayush/page/61","https://www.netmeds.com/non-prescriptions/ayush/page/62",
          "https://www.netmeds.com/non-prescriptions/ayush/page/63","https://www.netmeds.com/non-prescriptions/ayush/page/64",
          "https://www.netmeds.com/non-prescriptions/ayush/page/65","https://www.netmeds.com/non-prescriptions/ayush/page/66",
          "https://www.netmeds.com/non-prescriptions/ayush/page/67","https://www.netmeds.com/non-prescriptions/ayush/page/68",
          "https://www.netmeds.com/non-prescriptions/ayush/page/69","https://www.netmeds.com/non-prescriptions/ayush/page/70",
          "https://www.netmeds.com/non-prescriptions/ayush/page/71","https://www.netmeds.com/non-prescriptions/ayush/page/72",
          "https://www.netmeds.com/non-prescriptions/ayush/page/73","https://www.netmeds.com/non-prescriptions/ayush/page/73",
          "https://www.netmeds.com/non-prescriptions/ayush/page/74","https://www.netmeds.com/non-prescriptions/ayush/page/75",
          "https://www.netmeds.com/non-prescriptions/ayush/page/75","https://www.netmeds.com/non-prescriptions/ayush/page/76",
          "https://www.netmeds.com/non-prescriptions/ayush/page/77","https://www.netmeds.com/non-prescriptions/ayush/page/78",
          "https://www.netmeds.com/non-prescriptions/ayush/page/78","https://www.netmeds.com/non-prescriptions/ayush/page/79"]
    
    for links in link:
        html = requests.get(links) 
        #html = links.content
        soup = BeautifulSoup(html.text, 'html.parser')
        #print(soup)
        product=soup.find("div",class_="row product-list").find_all("div",class_="cat-item")
        for products in product:
            #print(products)
            product_name=products.find("a").find("span",class_="clsgetname").get_text(strip=True)
            product_price=products.find('span', attrs={'id':'final_price'}).get_text(strip=True).replace("Rs.","")
            product_owner=products.find("span",class_="drug-varients ellipsis").get_text(strip=True).replace("Mkt:","")
            #print(product_name,product_price,product_owner)
            ws4.append([product_name,product_price,product_owner])
    
    
except Exception as e:
    print(e)
        

wb.save(filename = 'Product Details.xlsx')
 

